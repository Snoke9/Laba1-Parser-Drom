from bs4 import BeautifulSoup
from openpyxl import load_workbook
import requests


def parse():
    car_name = []
    car_price = []
    car_cities = []
    car_links = []
    for i in range(1, 20):  # НА ПЕРВЫХ ~10 СТРАНИЦАХ НАХОДЯТСЯ ЗАКРЕПЛЕННЫЕ ОБЪЯВЛЕНИЯ. НОВЫЕ ОТОБРАЖАЮТСЯ ДАЛЬШЕ.
        url = f'https://auto.drom.ru/all/page{i}/'  # ПРОГА ИХ СОРТИРУЕТ, НО ДЛЯ БОЛЕЕ БЫСТРОЙ РАБОТЫ ЛУЧШЕ ПОСМОТРЕТЬ,
        page = requests.get(url)                    # ГДЕ СЕЙЧАС ОТОБРАЖАЮТСЯ НОВЫЕ ОБЪЯВЛЕНИЯ, И ОГРАНИЧИТЬ range
        soup = BeautifulSoup(page.text, "html.parser")                                # ПО НОМЕРАМ СТРАНИЦ
        names = soup.findAll('div', class_='css-16kqa8y e3f4v4l2')  # ИМЯ МАШИНЫ
        inaccuracy_counter = 0
        for name in names:
            if len(car_name) == 20:  # НУЖНОЕ КОЛИЧЕСТВО ОБЪЯВЛЕНИЙ; РЕГУЛЯТОР КОЛИЧЕСТВА
                break
            if name.find('div', class_='css-1ajc6qi e3f4v4l1'):  # ПРОВЕРКА НА НАМЕРЕННО ЗАКРЕПЛЕННОЕ ОБЪЯВЛЕНИЕ
                inaccuracy_counter += 1
                continue
            elif name.find('span'):  # ЗАПИСЬ НОВОГО ОБЪЯВЛЕНИЯ В СПИСОК
                carny = name.text
                car_name.append(carny)

        price = soup.findAll('span', class_='css-46itwz e162wx9x0')  # ЦЕНА МАШИНЫ
        for rubles in price:
            if len(car_price) == inaccuracy_counter + len(car_name):  # НУЖНОЕ КОЛИЧЕСТВО ЦЕННИКОВ
                break
            if rubles.find('span'):  # ЗАПИСЬ ЦЕННИКА В СПИСОК
                carp = rubles.text
                carp = carp.replace('\xa0', '')[:-1]
                car_price.append(carp)
        car_price = car_price[inaccuracy_counter:]  # СРЕЗ ЦЕННИКОВ ЗАКРЕПЛЕННЫХ ОБЪЯВЛЕНИЙ

        cities = soup.findAll('span', class_='css-1488ad e162wx9x0')  # ГОРОД МАШИНЫ
        for city in cities:
            if len(car_cities) == inaccuracy_counter + len(car_name):  # НУЖНОЕ КОЛИЧЕСТВО ГОРОДОВ
                break
            car_city = city.text
            car_cities.append(car_city)  # ЗАПИСЬ ГОРОДА В СПИСОК
        car_cities = car_cities[inaccuracy_counter:]  # СРЕЗ ГОРОДОВ ЗАКРЕПЛЕННЫХ ОБЪЯВЛЕНИЙ

        link = soup.findAll('a', class_='css-4zflqt e1huvdhj1')  # ПОЛУЧЕНИЕ ССЫЛКИ
        for links in link:
            if len(car_links) == inaccuracy_counter + len(car_name):  # НУЖНОЕ КОЛИЧЕСТВО ССЫЛОК
                break
            car_link = links.get('href')
            car_links.append(car_link)
        car_links = car_links[inaccuracy_counter:]  # СРЕЗ ССЫЛОК ЗАКРЕПЛЕННЫХ ОБЪЯВЛЕНИЙ

    car_years = [0] * len(car_name)
    for i in range(len(car_name)):  # СОЗДАНИЕ СПИСКА ГОДОВ ОТДЕЛЬНО ОТ ИМЕНИ
        car_years[i] = car_name[i][-4:]
        car_name[i] = car_name[i][:-6]
    return car_name, car_years, car_cities, car_price, car_links


def filing(info):
    excel_file = load_workbook('data_base.xlsx')
    excel_file_page1 = excel_file['about_cars']
    excel_file_page1.delete_rows(2, excel_file_page1.max_row)  # УДАЛЕНИЕ ПРЕДЫДУЩИХ ЗАПИСЕЙ ИЗ ФАЙЛА
    excel_file.save('data_base.xlsx')
    for data in range(len(info[0])):  # ЗАПИСЬ НОВЫХ ОБЪЯВЛЕНИЙ В EXCEL ФАЙЛ
        excel_file_page1.append(([info[0][data], int(info[1][data]), info[2][data], int(info[3][data]), info[4][data]]))
    excel_file.save('data_base.xlsx')
    excel_file.close()
